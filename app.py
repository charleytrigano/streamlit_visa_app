# Visa Manager - app.py
# Final corrected Streamlit application (single file)
# - Ensures every st.date_input gets a native datetime.date or None via _date_or_none_safe
# - All date_input calls use _date_or_none_safe(...) directly (no intermediate wrappers that might be shadowed)
# - Gestion form includes st.form_submit_button
# - Robust CSV/XLSX reading (semicolon supported), heuristic mapping
# - Allows negative Solde
#
# Requirements: pip install streamlit pandas openpyxl
# Run: streamlit run app.py

import os
import json
import re
from io import BytesIO
from datetime import date, datetime
from typing import Tuple, Dict, Any, List, Optional

import pandas as pd
import streamlit as st

# Optional plotting
try:
    import plotly.express as px
    HAS_PLOTLY = True
except Exception:
    px = None
    HAS_PLOTLY = False

# Optional openpyxl for XLSX formula exports
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

# -------------------------
# Configuration & constants
# -------------------------
APP_TITLE = "🛂 Visa Manager"
COLS_CLIENTS = [
    "ID_Client", "Dossier N", "Nom", "Date",
    "Categories", "Sous-categorie", "Visa",
    "Montant honoraires (US $)", "Autres frais (US $)",
    "Payé", "Solde", "Solde à percevoir (US $)",
    "Acompte 1", "Date Acompte 1",
    "Acompte 2", "Date Acompte 2", "Acompte 3", "Date Acompte 3", "Acompte 4", "Date Acompte 4",
    "Escrow",
    "RFE", "Dossiers envoyé", "Dossier approuvé",
    "Dossier refusé", "Dossier Annulé",
    "Date d'envoi",
    "Date de création", "Créé par", "Dernière modification", "Modifié par",
    "Commentaires"
]
MEMO_FILE = "_vmemory.json"
CACHE_CLIENTS = "_clients_cache.bin"
CACHE_VISA = "_visa_cache.bin"
SHEET_CLIENTS = "Clients"
SHEET_VISA = "Visa"
SID = "vmgr"
DEFAULT_START_CLIENT_ID = 13057

CURRENT_USER = "charleytrigano"

def skey(*parts: str) -> str:
    return f"{SID}_" + "_".join([p for p in parts if p])

# -------------------------
# Helpers: parsing/formatting
# -------------------------
def normalize_header_text(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s

def remove_accents(s: Any) -> str:
    if s is None:
        return ""
    s2 = str(s)
    replace_map = {
        "é":"e","è":"e","ê":"e","ë":"e",
        "à":"a","â":"a",
        "î":"i","ï":"i",
        "ô":"o","ö":"o",
        "ù":"u","û":"u","ü":"u",
        "ç":"c"
    }
    for k,v in replace_map.items():
        s2 = s2.replace(k, v)
    return s2

def canonical_key(s: Any) -> str:
    if s is None:
        return ""
    s2 = normalize_header_text(str(s)).lower()
    s2 = remove_accents(s2)
    s2 = re.sub(r"[^a-z0-9 ]", " ", s2)
    s2 = re.sub(r"\s+", " ", s2).strip()
    return s2

def money_to_float(x: Any) -> float:
    try:
        if pd.isna(x):
            return 0.0
    except Exception:
        pass
    try:
        s = str(x).strip()
        if s == "" or s.lower() in ("na","n/a","nan"):
            return 0.0
        s = s.replace("\u202f", "").replace("\xa0", "").replace(" ", "")
        s = re.sub(r"[^\d,.\-]", "", s)
        if s == "":
            return 0.0
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        else:
            if "," in s and s.count(",") == 1 and "." not in s:
                if len(s.split(",")[-1]) == 2:
                    s = s.replace(",", ".")
                else:
                    s = s.replace(",", "")
            else:
                s = s.replace(",", ".")
        return float(s)
    except Exception:
        try:
            return float(re.sub(r"[^0-9.\-]", "", str(x)))
        except Exception:
            return 0.0

def _to_num(x: Any) -> float:
    if isinstance(x, (int, float)):
        return float(x)
    return money_to_float(x)

def _fmt_money(v: Any) -> str:
    try:
        return "${:,.2f}".format(float(v))
    except Exception:
        return "$0.00"

def _date_or_none_safe(v: Any) -> Optional[date]:
    """
    Strict safe wrapper: always return a native datetime.date or None.
    Use this for all st.date_input value=... calls to avoid pandas types.
    """
    try:
        if v is None:
            return None
        if isinstance(v, date) and not isinstance(v, datetime):
            return v
        if isinstance(v, datetime):
            return v.date()
        # pandas Timestamp, numpy datetime64, string -> convert via pandas then build native date
        d = pd.to_datetime(v, errors="coerce")
        if pd.isna(d):
            return None
        # Ensure we return a Python datetime.date
        return date(int(d.year), int(d.month), int(d.day))
    except Exception:
        return None

# -------------------------
# Column heuristics & detectors
# -------------------------
COL_CANDIDATES = {
    "id client": "ID_Client", "idclient": "ID_Client",
    "dossier n": "Dossier N", "dossier": "Dossier N",
    "nom": "Nom", "date": "Date",
    "categories": "Categories", "categorie": "Categories",
    "sous categorie": "Sous-categorie", "sous-categorie": "Sous-categorie", "souscategorie": "Sous-categorie",
    "visa": "Visa",
    "montant": "Montant honoraires (US $)", "montant honoraires": "Montant honoraires (US $)",
    "autres frais": "Autres frais (US $)", "autresfrais": "Autres frais (US $)",
    "payé": "Payé", "paye": "Payé",
    "solde": "Solde",
    "solde a percevoir": "Solde à percevoir (US $)",
    "acompte 1": "Acompte 1", "acompte1": "Acompte 1",
    "date acompte 1": "Date Acompte 1",
    "acompte 2": "Acompte 2", "acompte2": "Acompte 2",
    "acompte 3": "Acompte 3", "acompte3": "Acompte 3",
    "acompte 4": "Acompte 4", "acompte4": "Acompte 4",
    "escrow": "Escrow",
    "dossier envoye": "Dossiers envoyé", "dossier approuve": "Dossier approuvé", "dossier refuse": "Dossier refusé",
    "rfe": "RFE", "commentaires": "Commentaires"
}

NUMERIC_TARGETS = [
    "Montant honoraires (US $)",
    "Autres frais (US $)",
    "Payé",
    "Solde",
    "Solde à percevoir (US $)",
    "Acompte 1",
    "Acompte 2",
    "Acompte 3",
    "Acompte 4"
]

def detect_acompte_columns(df: pd.DataFrame) -> List[str]:
    cols = []
    if df is None or df.empty:
        return cols
    for c in df.columns:
        k = canonical_key(c)
        if "acompte" in k:
            cols.append(c)
    def sort_key(name):
        m = re.search(r"(\d+)", name)
        return int(m.group(1)) if m else 999
    cols = sorted(cols, key=sort_key)
    return cols

def detect_montant_column(df: pd.DataFrame) -> Optional[str]:
    if df is None or df.empty:
        return None
    candidates = ["Montant honoraires (US $)", "Montant honoraires", "Montant", "Montant honoraires (USD)"]
    for c in candidates:
        if c in df.columns:
            return c
    for c in df.columns:
        k = canonical_key(c)
        if "montant" in k or "honorair" in k:
            return c
    return None

def detect_autres_column(df: pd.DataFrame) -> Optional[str]:
    if df is None or df.empty:
        return None
    candidates = ["Autres frais (US $)", "Autres frais", "Autres"]
    for c in candidates:
        if c in df.columns:
            return c
    for c in df.columns:
        k = canonical_key(c)
        if "autre" in k or "frais" in k:
            return c
    return None

def map_columns_heuristic(df: Any) -> Tuple[pd.DataFrame, Dict[str,str]]:
    if not isinstance(df, pd.DataFrame):
        return pd.DataFrame(), {}
    mapping: Dict[str,str] = {}
    for c in list(df.columns):
        key = canonical_key(c)
        mapped = None
        if key in COL_CANDIDATES:
            mapped = COL_CANDIDATES[key]
        else:
            for cand_key, std in sorted(COL_CANDIDATES.items(), key=lambda t: -len(t[0])):
                if cand_key in key:
                    mapped = std
                    break
        if mapped is None:
            mapped = normalize_header_text(c)
        mapping[c] = mapped
    new_names = {}
    seen = {}
    for orig, new in mapping.items():
        base = new
        cnt = seen.get(base, 0)
        if cnt:
            new_name = f"{base}_{cnt+1}"
            seen[base] = cnt+1
        else:
            new_name = base
            seen[base] = 1
        new_names[orig] = new_name
    try:
        df = df.rename(columns=new_names)
    except Exception:
        pass
    return df, new_names

def coerce_category_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    cols = list(df.columns)
    rename_map = {}
    def _ck(x): return canonical_key(str(x))
    for c in cols:
        k = _ck(c)
        if ("sous" in k and "categorie" in k) or ("souscategorie" in k):
            if "Sous-categorie" not in df.columns:
                rename_map[c] = "Sous-categorie"
        elif ("categorie" in k or "categories" in k) and "sous" not in k:
            if "Categories" not in df.columns:
                rename_map[c] = "Categories"
    if rename_map:
        try:
            df = df.rename(columns=rename_map)
        except Exception:
            pass
    return df

# -------------------------
# Visa helper
# -------------------------
DEFAULT_VISA_OPTIONS_BY_CAT_SUB: Dict[Tuple[str,str], List[str]] = {}
visa_sub_options_map: Dict[str, List[str]] = {}
visa_map: Dict[str, List[str]] = {}
visa_map_norm: Dict[str, List[str]] = {}
visa_categories: List[str] = []

def get_visa_options(cat: Optional[str], sub: Optional[str]) -> List[str]:
    global visa_sub_options_map, visa_map, visa_map_norm, DEFAULT_VISA_OPTIONS_BY_CAT_SUB
    try:
        if sub:
            ksub = canonical_key(sub)
            if isinstance(visa_sub_options_map, dict) and ksub in visa_sub_options_map:
                opts = visa_sub_options_map.get(ksub, [])
                if opts:
                    return opts[:]
    except Exception:
        pass
    try:
        if cat:
            kcat = canonical_key(cat)
            if isinstance(visa_map_norm, dict) and kcat in visa_map_norm:
                return visa_map_norm.get(kcat, [])[:]
    except Exception:
        pass
    try:
        if cat and sub:
            key = (canonical_key(cat), canonical_key(sub))
            if key in DEFAULT_VISA_OPTIONS_BY_CAT_SUB:
                return DEFAULT_VISA_OPTIONS_BY_CAT_SUB[key][:]
    except Exception:
        pass
    try:
        if sub:
            ksub = canonical_key(sub)
            for (kcat, ksub_k), opts in DEFAULT_VISA_OPTIONS_BY_CAT_SUB.items():
                if ksub_k == ksub:
                    return opts[:]
    except Exception:
        pass
    return []

# -------------------------
# I/O helpers (robust)
# -------------------------
def try_read_excel_from_bytes(b: bytes, sheet_name: Optional[str] = None) -> Optional[pd.DataFrame]:
    bio = BytesIO(b)
    try:
        xls = pd.ExcelFile(bio, engine="openpyxl")
        sheets = xls.sheet_names
        if sheet_name and sheet_name in sheets:
            return pd.read_excel(BytesIO(b), sheet_name=sheet_name, engine="openpyxl")
        for cand in [SHEET_CLIENTS, SHEET_VISA, "Sheet1"]:
            if cand in sheets:
                try:
                    return pd.read_excel(BytesIO(b), sheet_name=cand, engine="openpyxl")
                except Exception:
                    continue
        return pd.read_excel(BytesIO(b), sheet_name=0, engine="openpyxl")
    except Exception:
        return None

def read_any_table(src: Any, sheet: Optional[str] = None, debug_prefix: str = "") -> Optional[pd.DataFrame]:
    def _log(msg: str):
        try:
            st.sidebar.info(f"{debug_prefix}{msg}")
        except Exception:
            pass
    if src is None:
        _log("read_any_table: src is None")
        return None
    try:
        if isinstance(src, (bytes, bytearray)):
            df = try_read_excel_from_bytes(bytes(src), sheet)
            if df is not None:
                return df
            for sep in [";", ","]:
                for enc in ["utf-8", "latin-1", "cp1252"]:
                    try:
                        return pd.read_csv(BytesIO(src), sep=sep, encoding=enc, on_bad_lines="skip")
                    except Exception:
                        continue
            return None
        if isinstance(src, BytesIO):
            b = src.getvalue()
            df = try_read_excel_from_bytes(b, sheet)
            if df is not None:
                return df
            for sep in [";", ","]:
                for enc in ["utf-8", "latin-1", "cp1252"]:
                    try:
                        return pd.read_csv(BytesIO(b), sep=sep, encoding=enc, on_bad_lines="skip")
                    except Exception:
                        continue
            return None
        if hasattr(src, "read") and hasattr(src, "name"):
            try:
                data = src.getvalue()
            except Exception:
                try:
                    src.seek(0); data = src.read()
                except Exception:
                    data = None
            if data:
                df = try_read_excel_from_bytes(data, sheet)
                if df is not None:
                    return df
                for sep in [";", ","]:
                    for enc in ["utf-8", "latin-1", "cp1252"]:
                        try:
                            return pd.read_csv(BytesIO(data), sep=sep, encoding=enc, on_bad_lines="skip")
                        except Exception:
                            continue
            return None
        if isinstance(src, (str, os.PathLike)):
            p = str(src)
            if not os.path.exists(p):
                _log(f"path does not exist: {p}")
                return None
            if p.lower().endswith(".csv"):
                for sep in [";", ","]:
                    for enc in ["utf-8", "latin-1", "cp1252"]:
                        try:
                            return pd.read_csv(p, sep=sep, encoding=enc, on_bad_lines="skip")
                        except Exception:
                            continue
                return None
            else:
                try:
                    return pd.read_excel(p, sheet_name=sheet or 0, engine="openpyxl")
                except Exception:
                    return None
    except Exception as e:
        _log(f"read_any_table exception: {e}")
        return None
    _log("read_any_table: unsupported src type")
    return None

# -------------------------
# Ensure & normalize
# -------------------------
def _ensure_columns(df: Any, cols: List[str]) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame):
        df = pd.DataFrame()
    out = df.copy()
    for c in cols:
        if c not in out.columns:
            if c in ["Payé", "Solde", "Solde à percevoir (US $)", "Montant honoraires (US $)", "Autres frais (US $)",
                     "Acompte 1", "Acompte 2", "Acompte 3", "Acompte 4"]:
                out[c] = 0.0
            elif c in ["RFE", "Dossiers envoyé", "Dossier approuvé", "Dossier refusé", "Dossier Annulé"]:
                out[c] = 0
            elif c in ["Date de création", "Dernière modification", "Date", "Date Acompte 1", "Date Acompte 2", "Date Acompte 3", "Date Acompte 4", "Date d'envoi"]:
                out[c] = pd.NaT
            elif c == "Escrow":
                out[c] = 0
            else:
                out[c] = ""
    try:
        return out[cols]
    except Exception:
        safe = pd.DataFrame(columns=cols)
        for c in cols:
            if c in out.columns:
                safe[c] = out[c]
            else:
                if c in ["Payé", "Solde", "Solde à percevoir (US $)", "Montant honoraires (US $)", "Autres frais (US $)",
                         "Acompte 1", "Acompte 2", "Acompte 3", "Acompte 4"]:
                    safe[c] = 0.0
                elif c in ["Date de création", "Dernière modification", "Date", "Date Acompte 1", "Date Acompte 2", "Date Acompte 3", "Date Acompte 4", "Date d'envoi"]:
                    safe[c] = pd.NaT
                elif c == "Escrow":
                    safe[c] = 0
                else:
                    safe[c] = ""
        return safe

def normalize_clients_for_live(df_clients_raw: Any) -> pd.DataFrame:
    if not isinstance(df_clients_raw, pd.DataFrame):
        maybe_df = read_any_table(df_clients_raw, sheet=None, debug_prefix="[normalize] ")
        df_clients_raw = maybe_df if isinstance(maybe_df, pd.DataFrame) else pd.DataFrame()
    df_mapped, _ = map_columns_heuristic(df_clients_raw)
    for dtc in ["Date","Date de création","Dernière modification","Date Acompte 1","Date Acompte 2","Date Acompte 3","Date Acompte 4","Date d'envoi"]:
        if dtc in df_mapped.columns:
            try:
                df_mapped[dtc] = pd.to_datetime(df_mapped[dtc], dayfirst=True, errors="coerce")
            except Exception:
                pass
    df = _ensure_columns(df_mapped, COLS_CLIENTS)
    for col in NUMERIC_TARGETS:
        if col in df.columns:
            try:
                df[col] = df[col].apply(lambda x: _to_num(x))
            except Exception:
                df[col] = 0.0
    for acc in ["Acompte 1","Acompte 2","Acompte 3","Acompte 4"]:
        if acc not in df.columns:
            df[acc] = 0.0
    acomptes_cols = detect_acompte_columns(df)
    if acomptes_cols:
        try:
            df["Payé"] = df[acomptes_cols].fillna(0).apply(lambda row: sum([_to_num(row[c]) for c in acomptes_cols]), axis=1)
        except Exception:
            df["Payé"] = df.get("Payé", 0).apply(lambda x: _to_num(x))
    try:
        montant_col = detect_montant_column(df) or "Montant honoraires (US $)"
        autres_col = detect_autres_column(df) or "Autres frais (US $)"
        df[montant_col] = df.get(montant_col, 0).apply(lambda x: _to_num(x))
        df[autres_col] = df.get(autres_col, 0).apply(lambda x: _to_num(x))
        df["Payé"] = df.get("Payé", 0).apply(lambda x: _to_num(x))
        df["Solde"] = df[montant_col] + df[autres_col] - df["Payé"]
        df["Solde à percevoir (US $)"] = df["Solde"].copy()
    except Exception:
        df["Solde"] = df.get("Solde", 0).apply(lambda x: _to_num(x))
        df["Solde à percevoir (US $)"] = df.get("Solde à percevoir (US $)", 0).apply(lambda x: _to_num(x))
    for f in ["RFE","Dossiers envoyé","Dossier approuvé","Dossier refusé","Dossier Annulé"]:
        if f not in df.columns:
            df[f] = 0
    if "Escrow" not in df.columns:
        df["Escrow"] = 0
    for c in ["Nom","Categories","Sous-catégorie","Visa","Commentaires","Créé par","Modifié par"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)
    try:
        dser = pd.to_datetime(df["Date"], errors="coerce")
        df["_Année_"] = dser.dt.year.fillna(0).astype(int)
        df["_MoisNum_"] = dser.dt.month.fillna(0).astype(int)
        df["Mois"] = df["_MoisNum_"].apply(lambda m: f"{int(m):02d}" if pd.notna(m) and m>0 else "")
    except Exception:
        df["_Année_"] = 0; df["_MoisNum_"] = 0; df["Mois"] = ""
    return df

def recalc_payments_and_solde(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    acomptes = detect_acompte_columns(out)
    if not acomptes:
        for acc in ["Acompte 1","Acompte 2","Acompte 3","Acompte 4"]:
            if acc not in out.columns:
                out[acc] = 0.0
        acomptes = detect_acompte_columns(out)
    for c in acomptes:
        try:
            out[c] = out[c].apply(lambda x: _to_num(x) if not isinstance(x,(int,float)) else float(x))
        except Exception:
            out[c] = out[c].apply(lambda x: 0.0)
    montant_col = detect_montant_column(out) or "Montant honoraires (US $)"
    autres_col = detect_autres_column(out) or "Autres frais (US $)"
    for c in [montant_col, autres_col]:
        if c not in out.columns:
            out[c] = 0.0
        else:
            try:
                out[c] = out[c].apply(lambda x: _to_num(x) if not isinstance(x,(int,float)) else float(x))
            except Exception:
                out[c] = out[c].apply(lambda x: 0.0)
    try:
        out["Payé"] = out[acomptes].sum(axis=1).astype(float) if acomptes else out.get("Payé",0).apply(lambda x: _to_num(x))
    except Exception:
        out["Payé"] = out.get("Payé",0).apply(lambda x: _to_num(x))
    try:
        out["Solde"] = out[montant_col] + out[autres_col] - out["Payé"]
        out["Solde à percevoir (US $)"] = out["Solde"].copy()
        out["Solde"] = out["Solde"].astype(float)
    except Exception:
        out["Solde"] = out.get("Solde",0).apply(lambda x: _to_num(x))
    if "Escrow" in out.columns:
        try:
            out["Escrow"] = out["Escrow"].apply(lambda x: 1 if str(x).strip().lower() in ("1","true","t","yes","oui","y","x") else (1 if _to_num(x) == 1 else 0))
        except Exception:
            out["Escrow"] = out["Escrow"].apply(lambda x: 1 if str(x).strip().lower() in ("1","true","t","yes","oui","y","x") else 0)
    return out

def get_next_dossier_numeric(df: pd.DataFrame) -> int:
    if df is None or df.empty:
        return DEFAULT_START_CLIENT_ID
    vals = df.get("Dossier N", pd.Series([], dtype="object"))
    nums = []
    for v in vals.dropna().astype(str):
        m = re.search(r"(\d+)", v)
        if m:
            try:
                nums.append(int(m.group(1)))
            except Exception:
                pass
    if not nums:
        return DEFAULT_START_CLIENT_ID
    mx = max(nums)
    return max(DEFAULT_START_CLIENT_ID, mx) + 1

def make_id_client_datebased(df: pd.DataFrame) -> str:
    seq = get_next_dossier_numeric(df)
    datepart = datetime.now().strftime("%Y%m%d")
    return f"{datepart}-{seq}"

def ensure_flag_columns(df: pd.DataFrame, flags: List[str]) -> None:
    for f in flags:
        if f not in df.columns:
            df[f] = 0

DEFAULT_FLAGS = ["RFE", "Dossiers envoyé", "Dossier approuvé", "Dossier refusé", "Dossier Annulé"]

# -------------------------
# UI bootstrap & upload handling
# -------------------------
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)

st.sidebar.header("📂 Fichiers")
last_clients_path = ""
last_visa_path = ""
try:
    if os.path.exists(MEMO_FILE):
        with open(MEMO_FILE, "r", encoding="utf-8") as f:
            d = json.load(f)
            last_clients_path = d.get("clients","")
            last_visa_path = d.get("visa","")
except Exception:
    pass

up_clients = st.sidebar.file_uploader("Clients (xlsx/xls/csv)", type=["xlsx","xls","csv"], key=skey("up_clients"))
up_visa = st.sidebar.file_uploader("Visa (xlsx/xls/csv)", type=["xlsx","xls","csv"], key=skey("up_visa"))
clients_path_in = st.sidebar.text_input("ou chemin local Clients (optionnel)", value=last_clients_path or "", key=skey("cli_path"))
visa_path_in = st.sidebar.text_input("ou chemin local Visa (optionnel)", value=last_visa_path or "", key=skey("vis_path"))

if st.sidebar.button("📥 Sauvegarder chemins", key=skey("btn_save_paths")):
    try:
        with open(MEMO_FILE, "w", encoding="utf-8") as f:
            json.dump({"clients": clients_path_in or "", "visa": visa_path_in or ""}, f, ensure_ascii=False, indent=2)
        st.sidebar.success("Chemins sauvegardés.")
    except Exception:
        st.sidebar.error("Impossible de sauvegarder les chemins.")

# persist uploaded bytes
clients_bytes = None
visa_bytes = None
if up_clients is not None:
    try:
        clients_bytes = up_clients.getvalue()
        with open(CACHE_CLIENTS, "wb") as f:
            f.write(clients_bytes)
    except Exception:
        pass
if up_visa is not None:
    try:
        visa_bytes = up_visa.getvalue()
        with open(CACHE_VISA, "wb") as f:
            f.write(visa_bytes)
    except Exception:
        pass

if clients_bytes is not None:
    clients_src_for_read = BytesIO(clients_bytes)
elif clients_path_in:
    clients_src_for_read = clients_path_in
elif os.path.exists(CACHE_CLIENTS):
    try:
        clients_bytes = open(CACHE_CLIENTS, "rb").read()
        clients_src_for_read = BytesIO(clients_bytes)
    except Exception:
        clients_src_for_read = None
else:
    clients_src_for_read = None

if visa_bytes is not None:
    visa_src_for_read = BytesIO(visa_bytes)
elif visa_path_in:
    visa_src_for_read = visa_path_in
elif os.path.exists(CACHE_VISA):
    try:
        visa_bytes = open(CACHE_VISA, "rb").read()
        visa_src_for_read = BytesIO(visa_bytes)
    except Exception:
        visa_src_for_read = None
else:
    visa_src_for_read = None

# -------------------------
# Read raw tables
# -------------------------
df_clients_raw = None
df_visa_raw = None
try:
    df_clients_raw = read_any_table(clients_src_for_read, sheet=SHEET_CLIENTS, debug_prefix="[Clients] ")
except Exception:
    df_clients_raw = None
if df_clients_raw is None and clients_src_for_read is not None:
    df_clients_raw = read_any_table(clients_src_for_read, sheet=None, debug_prefix="[Clients fallback] ")
if df_clients_raw is None:
    df_clients_raw = pd.DataFrame()

try:
    df_visa_raw = read_any_table(visa_src_for_read, sheet=SHEET_VISA, debug_prefix="[Visa] ")
except Exception:
    df_visa_raw = None
if df_visa_raw is None and visa_src_for_read is not None:
    df_visa_raw = read_any_table(visa_src_for_read, sheet=None, debug_prefix="[Visa fallback] ")
if df_visa_raw is None:
    df_visa_raw = pd.DataFrame()

# sanitize visa raw
if isinstance(df_visa_raw, pd.DataFrame) and not df_visa_raw.empty:
    try:
        df_visa_raw = df_visa_raw.fillna("")
        for c in df_visa_raw.columns:
            try:
                df_visa_raw[c] = df_visa_raw[c].astype(str).str.strip()
            except Exception:
                pass
    except Exception:
        pass

# build visa maps
visa_map = {}; visa_map_norm = {}; visa_categories = []; visa_sub_options_map = {}
if isinstance(df_visa_raw, pd.DataFrame) and not df_visa_raw.empty:
    df_visa_mapped, _ = map_columns_heuristic(df_visa_raw)
    try:
        df_visa_mapped = coerce_category_columns(df_visa_mapped)
    except Exception:
        pass
    raw_vm = {}
    try:
        for _, row in df_visa_mapped.iterrows():
            cat = str(row.get("Categories","")).strip()
            sub = str(row.get("Sous-categorie","")).strip()
            if not cat:
                continue
            raw_vm.setdefault(cat, [])
            if sub and sub not in raw_vm[cat]:
                raw_vm[cat].append(sub)
    except Exception:
        raw_vm = {}
    raw_vm = {k: [s for s in v if s and str(s).strip().lower() != "nan"] for k, v in raw_vm.items()}
    visa_map = {k.strip(): [s.strip() for s in v] for k, v in raw_vm.items()}
    visa_map_norm = {canonical_key(k): v for k, v in visa_map.items()}
    visa_categories = sorted(list(visa_map.keys()))
    visa_sub_options_map = {}
    try:
        cols_to_skip = set(["Categories","Categorie","Sous-categorie"])
        cols_to_check = [c for c in df_visa_mapped.columns if c not in cols_to_skip]
        for _, row in df_visa_mapped.iterrows():
            sub = str(row.get("Sous-categorie","")).strip()
            if not sub:
                continue
            key = canonical_key(sub)
            for col in cols_to_check:
                val = row.get(col,"")
                truthy = False
                if pd.isna(val):
                    truthy = False
                else:
                    sval = str(val).strip().lower()
                    if sval in ("1","x","t","true","oui","yes","y"):
                        truthy = True
                    else:
                        try:
                            if float(sval) == 1.0:
                                truthy = True
                        except Exception:
                            truthy = False
                if truthy:
                    visa_sub_options_map.setdefault(key, [])
                    if col not in visa_sub_options_map[key]:
                        visa_sub_options_map[key].append(col)
    except Exception:
        visa_sub_options_map = {}

globals()['visa_map'] = visa_map
globals()['visa_map_norm'] = visa_map_norm
globals()['visa_categories'] = visa_categories
globals()['visa_sub_options_map'] = visa_sub_options_map

# -------------------------
# Build live DF in session
# -------------------------
df_all = normalize_clients_for_live(df_clients_raw)
df_all = recalc_payments_and_solde(df_all)
DF_LIVE_KEY = skey("df_live")
if isinstance(df_all, pd.DataFrame) and not df_all.empty:
    st.session_state[DF_LIVE_KEY] = df_all.copy()
else:
    if DF_LIVE_KEY not in st.session_state or st.session_state[DF_LIVE_KEY] is None:
        st.session_state[DF_LIVE_KEY] = pd.DataFrame(columns=COLS_CLIENTS)

def _get_df_live() -> pd.DataFrame:
    return st.session_state[DF_LIVE_KEY].copy()

def _set_df_live(df: pd.DataFrame) -> None:
    st.session_state[DF_LIVE_KEY] = df.copy()

# small helpers
def unique_nonempty(series):
    try:
        vals = series.dropna().astype(str).tolist()
    except Exception:
        vals = []
    out = []
    for v in vals:
        s = str(v).strip()
        if s == "" or s.lower() == "nan":
            continue
        out.append(s)
    return sorted(list(dict.fromkeys(out)))

def kpi_html(label: str, value: str, sub: str = "") -> str:
    html = f"""
    <div style="border:1px solid rgba(255,255,255,0.04); border-radius:6px; padding:8px 10px; margin:6px 4px;">
      <div style="font-size:12px; color:#666;">{label}</div>
      <div style="font-size:18px; font-weight:700; margin-top:4px;">{value}</div>
      <div style="font-size:11px; color:#888; margin-top:4px;">{sub}</div>
    </div>
    """
    return html

# -------------------------
# Tabs UI: Files / Dashboard / Analyses / Add / Gestion / Export
# -------------------------
tabs = st.tabs(["📄 Fichiers","📊 Dashboard","📈 Analyses","➕ Ajouter","✏️ / 🗑️ Gestion","💾 Export"])

# ---- Files tab ----
with tabs[0]:
    st.header("📂 Fichiers")
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Clients")
        if up_clients is not None:
            st.text(f"Upload: {getattr(up_clients,'name','')}")
        elif isinstance(clients_src_for_read, str) and clients_src_for_read:
            st.text(f"Chargé depuis: {clients_src_for_read}")
        elif os.path.exists(CACHE_CLIENTS):
            st.text("Chargé depuis le cache local")
        if df_clients_raw is None or df_clients_raw.empty:
            st.warning("Aucun fichier Clients detecté.")
        else:
            st.success(f"Clients lus: {df_clients_raw.shape[0]} lignes")
            try:
                max_preview = 100
                if df_clients_raw.shape[0] <= max_preview:
                    st.dataframe(df_clients_raw.reset_index(drop=True), use_container_width=True, height=360)
                else:
                    st.dataframe(df_clients_raw.head(100).reset_index(drop=True), use_container_width=True, height=360)
                    if st.button("Afficher tout (peut être lent)"):
                        st.dataframe(df_clients_raw.reset_index(drop=True), use_container_width=True, height=600)
            except Exception:
                st.write(df_clients_raw.head(8))
    with c2:
        st.subheader("Visa")
        if up_visa is not None:
            st.text(f"Upload: {getattr(up_visa,'name','')}")
        elif isinstance(visa_src_for_read, str) and visa_src_for_read:
            st.text(f"Chargé depuis: {visa_src_for_read}")
        elif os.path.exists(CACHE_VISA):
            st.text("Chargé depuis le cache local")
        if df_visa_raw is None or df_visa_raw.empty:
            st.warning("Aucun fichier Visa detecté.")
        else:
            st.success(f"Visa lu: {df_visa_raw.shape[0]} lignes, {df_visa_raw.shape[1]} colonnes")
            try:
                st.dataframe(df_visa_raw.reset_index(drop=True), use_container_width=True, height=360)
            except Exception:
                st.write(df_visa_raw.head(8))
    st.markdown("---")
    col_a, col_b = st.columns([1,1])
    with col_a:
        if st.button("Réinitialiser mémoire (recharger)"):
            df_all2 = normalize_clients_for_live(df_clients_raw)
            df_all2 = recalc_payments_and_solde(df_all2)
            _set_df_live(df_all2)
            st.success("Mémoire réinitialisée.")
            try:
                st.experimental_rerun()
            except Exception:
                pass
    with col_b:
        if st.button("Actualiser la lecture"):
            try:
                st.experimental_rerun()
            except Exception:
                pass

# ---- Dashboard tab ----
with tabs[1]:
    st.subheader("📊 Dashboard (totaux et diagnostics)")
    df_live_view = recalc_payments_and_solde(_get_df_live())
    if df_live_view is None or df_live_view.empty:
        st.info("Aucune donnée en mémoire. Vérifiez l'onglet Fichiers et chargez le CSV correctement.")
    else:
        cats = unique_nonempty(df_live_view["Categories"]) if "Categories" in df_live_view.columns else []
        subs = unique_nonempty(df_live_view["Sous-catégorie"]) if "Sous-catégorie" in df_live_view.columns else []
        visas = unique_nonempty(df_live_view["Visa"]) if "Visa" in df_live_view.columns else []
        years = []
        if "_Année_" in df_live_view.columns:
            try:
                years = sorted([int(y) for y in pd.to_numeric(df_live_view["_Année_"], errors="coerce").dropna().unique().astype(int).tolist()])
            except Exception:
                years = []
        f1, f2, f3, f4 = st.columns([1,1,1,1])
        sel_cat = f1.selectbox("Catégorie", options=[""]+cats, index=0, key=skey("dash","cat"))
        sel_sub = f2.selectbox("Sous-catégorie", options=[""]+subs, index=0, key=skey("dash","sub"))
        sel_visa = f3.selectbox("Visa", options=[""]+visas, index=0, key=skey("dash","visa"))
        year_options = ["Toutes les années"] + [str(y) for y in years]
        sel_year = f4.selectbox("Année", options=year_options, index=0, key=skey("dash","year"))
        view = df_live_view.copy()
        if sel_cat:
            view = view[view["Categories"].astype(str) == sel_cat]
        if sel_sub:
            view = view[view["Sous-catégorie"].astype(str) == sel_sub]
        if sel_visa:
            view = view[view["Visa"].astype(str) == sel_visa]
        if sel_year and sel_year != "Toutes les années":
            view = view[view["_Année_"].astype(str) == sel_year]
        view = recalc_payments_and_solde(view)
        def safe_num(x):
            try:
                return float(_to_num(x))
            except Exception:
                return 0.0
        montant_col = detect_montant_column(view) or "Montant honoraires (US $)"
        autres_col = detect_autres_column(view) or "Autres frais (US $)"
        acomptes_cols = detect_acompte_columns(view)
        view["_Montant_num_"] = view.get(montant_col, 0).apply(safe_num)
        view["_Autres_num_"] = view.get(autres_col, 0).apply(safe_num)
        total_acomptes_sum = 0.0
        if acomptes_cols:
            for c in acomptes_cols:
                view[f"_ac_{c}"] = view.get(c, 0).apply(safe_num)
                total_acomptes_sum += float(view[f"_ac_{c}"].sum())
        total_honoraires = float(view["_Montant_num_"].sum())
        total_autres = float(view["_Autres_num_"].sum())
        total_paye = float(total_acomptes_sum)
        canonical_solde_sum = float(total_honoraires + total_autres - total_paye)
        cols_k = st.columns(4)
        cols_k[0].markdown(kpi_html("Dossiers (vue)", f"{len(view):,}"), unsafe_allow_html=True)
        cols_k[1].markdown(kpi_html("Montant honoraires", _fmt_money(total_honoraires)), unsafe_allow_html=True)
        cols_k[2].markdown(kpi_html("Autres frais", _fmt_money(total_autres)), unsafe_allow_html=True)
        cols_k[3].markdown(kpi_html("Total facturé (recalc)", _fmt_money(total_honoraires + total_autres)), unsafe_allow_html=True)
        st.markdown("---")
        cols_k2 = st.columns(2)
        cols_k2[0].markdown(kpi_html("Montant payé (somme acomptes)", _fmt_money(total_paye)), unsafe_allow_html=True)
        cols_k2[1].markdown(kpi_html("Solde total (recalc)", _fmt_money(canonical_solde_sum)), unsafe_allow_html=True)
        st.markdown("### Détails — clients correspondant aux filtres")
        display_df = view.copy()
        if "Date" in display_df.columns:
            try:
                display_df["Date"] = pd.to_datetime(display_df["Date"], errors="coerce").dt.date.astype(str)
            except Exception:
                display_df["Date"] = display_df["Date"].astype(str)
        for dtc in ["Date Acompte 1", "Date Acompte 2", "Date d'envoi", "Date de création", "Dernière modification"]:
            if dtc in display_df.columns:
                try:
                    display_df[dtc] = pd.to_datetime(display_df[dtc], errors="coerce").dt.strftime("%Y-%m-%d")
                except Exception:
                    display_df[dtc] = display_df[dtc].astype(str)
        money_cols = [montant_col, autres_col, "Payé","Solde","Solde à percevoir (US $)"] + acomptes_cols
        for mc in money_cols:
            if mc in display_df.columns:
                try:
                    display_df[mc] = display_df[mc].apply(lambda x: _fmt_money(_to_num(x)))
                except Exception:
                    display_df[mc] = display_df[mc].astype(str)
        try:
            st.dataframe(display_df.reset_index(drop=True), use_container_width=True, height=360)
        except Exception:
            st.write("Impossible d'afficher la liste des clients (trop volumineuse). Utilisez l'export.")

# ---- Analyses tab ----
with tabs[2]:
    st.subheader("📈 Analyses")
    st.info("Graphiques et analyses basiques.")
    df_ = _get_df_live()
    if isinstance(df_, pd.DataFrame) and not df_.empty and "Categories" in df_.columns:
        cat_counts = df_["Categories"].value_counts().rename_axis("Categorie").reset_index(name="Nombre")
        if HAS_PLOTLY and px is not None:
            fig = px.pie(cat_counts, names="Categorie", values="Nombre", hole=0.4, title="Répartition par catégorie")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.bar_chart(cat_counts.set_index("Categorie")["Nombre"])

# ---- Add tab ----
with tabs[3]:
    st.subheader("➕ Ajouter un nouveau client")
    df_live = _get_df_live()
    next_dossier_num = get_next_dossier_numeric(df_live)
    next_dossier = str(next_dossier_num)
    next_id_client = make_id_client_datebased(df_live)
    st.markdown(f"**ID_Client (auto)**: {next_id_client}")
    st.markdown(f"**Dossier N (auto)**: {next_dossier}")
    add_date = st.date_input("Date (événement)", value=date.today(), key=skey("addtab","date"))
    add_nom = st.text_input("Nom du client", value="", placeholder="Nom complet du client", key=skey("addtab","nom"))
    if visa_categories:
        categories_options = visa_categories
    else:
        if "Categories" in df_live.columns:
            cats_series = df_live["Categories"].dropna().astype(str).apply(lambda s: s.strip())
            categories_options = sorted([c for c in dict.fromkeys(cats_series) if c and c.lower() != "nan"])
        else:
            categories_options = []
    r3c1, r3c2, r3c3 = st.columns([1.2,1.6,1.6])
    with r3c1:
        categories_local = [""] + [c.strip() for c in categories_options]
        add_cat = st.selectbox("Catégorie", options=categories_local, index=0, key=skey("addtab","cat"))
    with r3c2:
        add_sub_options = []
        if isinstance(add_cat, str) and add_cat.strip():
            cat_key = canonical_key(add_cat)
            if cat_key in visa_map_norm:
                add_sub_options = visa_map_norm.get(cat_key, [])[:]
            else:
                if add_cat in visa_map:
                    add_sub_options = visa_map.get(add_cat, [])[:]
        if not add_sub_options:
            try:
                add_sub_options = sorted({str(x).strip() for x in df_live["Sous-catégorie"].dropna().astype(str).tolist()})
            except Exception:
                add_sub_options = []
        add_sub = st.selectbox("Sous-catégorie", options=[""] + add_sub_options, index=0, key=skey("addtab","sub"))
    with r3c3:
        specific_options = get_visa_options(add_cat, add_sub)
        if specific_options:
            add_visa = st.selectbox("Visa (options)", options=[""] + specific_options, index=0, key=skey("addtab","visa"))
        else:
            add_visa = st.text_input("Visa", value="", key=skey("addtab","visa"))
    r4c1, r4c2, r4c3, r4c4 = st.columns([1.2,1.0,1.0,0.6])
    with r4c1:
        add_montant = st.text_input("Montant honoraires (US $)", value="0", key=skey("addtab","montant"))
    with r4c2:
        a1 = st.text_input("Acompte 1", value="0", key=skey("addtab","ac1"))
    with r4c3:
        a1_date = st.date_input("Date Acompte 1", value=None, key=skey("addtab","ac1_date"))
    with r4c4:
        escrow_checked = st.checkbox("Escrow", value=False, key=skey("addtab","escrow"))
    try:
        montant_val = money_to_float(add_montant)
        autres_val = 0.0
        paid_val = money_to_float(a1)
        solde_val = montant_val + autres_val - paid_val
        solde_display = _fmt_money(solde_val)
    except Exception:
        solde_display = _fmt_money(0)
    st.markdown(f"**Solde**: {solde_display}")
    add_comments = st.text_area("Commentaires", value="", key=skey("addtab","comments"))
    if st.button("Ajouter", key=skey("addtab","btn_add")):
        try:
            new_row = {c: "" for c in df_live.columns}
            new_row["ID_Client"] = next_id_client
            new_row["Dossier N"] = next_dossier
            new_row["Nom"] = add_nom
            new_row["Date"] = pd.to_datetime(add_date)
            new_row["Categories"] = add_cat.strip() if isinstance(add_cat, str) else add_cat
            new_row["Sous-categorie"] = add_sub.strip() if isinstance(add_sub, str) else add_sub
            new_row["Visa"] = add_visa
            new_row["Montant honoraires (US $)"] = money_to_float(add_montant)
            new_row["Autres frais (US $)"] = 0.0
            new_row["Acompte 1"] = money_to_float(a1)
            new_row["Date Acompte 1"] = pd.to_datetime(a1_date) if a1_date else pd.NaT
            new_row["Acompte 2"] = 0.0
            new_row["Acompte 3"] = 0.0
            new_row["Acompte 4"] = 0.0
            new_row["Date Acompte 2"] = pd.NaT
            new_row["Date Acompte 3"] = pd.NaT
            new_row["Date Acompte 4"] = pd.NaT
            new_row["Escrow"] = 1 if escrow_checked else 0
            paid_sum = new_row["Acompte 1"] + new_row["Acompte 2"] + new_row["Acompte 3"] + new_row["Acompte 4"]
            new_row["Payé"] = paid_sum
            new_row["Solde"] = new_row["Montant honoraires (US $)"] + new_row["Autres frais (US $)"] - paid_sum
            new_row["Solde à percevoir (US $)"] = new_row["Solde"]
            now = datetime.now()
            new_row["Date de création"] = now
            new_row["Créé par"] = CURRENT_USER
            new_row["Dernière modification"] = now
            new_row["Modifié par"] = CURRENT_USER
            new_row["Commentaires"] = add_comments
            flags_to_create = DEFAULT_FLAGS
            ensure_flag_columns(df_live, flags_to_create)
            for opt in flags_to_create:
                new_row[opt] = 0
            new_row["Date d'envoi"] = pd.NaT
            df_live = pd.concat([df_live, pd.DataFrame([new_row])], ignore_index=True)
            df_live = recalc_payments_and_solde(df_live)
            _set_df_live(df_live)
            st.success(f"Dossier ajouté : ID_Client {next_id_client} — Dossier N {next_dossier}")
        except Exception as e:
            st.error(f"Erreur ajout: {e}")

# ---- Gestion tab ----
with tabs[4]:
    st.subheader("✏️ / 🗑️ Gestion — Modifier / Supprimer")
    df_live = _get_df_live()
    for c in COLS_CLIENTS:
        if c not in df_live.columns:
            if c in ["Date Acompte 2","Date Acompte 3","Date Acompte 4","Date d'envoi","Date de création","Dernière modification"]:
                df_live[c] = pd.NaT
            else:
                df_live[c] = "" if c not in NUMERIC_TARGETS else 0.0

    if df_live is None or df_live.empty:
        st.info("Aucun dossier à modifier ou supprimer.")
    else:
        choices = [f"{i} | {df_live.at[i,'Dossier N'] if 'Dossier N' in df_live.columns else ''} | {df_live.at[i,'Nom'] if 'Nom' in df_live.columns else ''}" for i in range(len(df_live))]
        sel = st.selectbox("Sélectionner ligne à modifier", options=[""]+choices, key=skey("edit","select"))
        if sel:
            idx = int(sel.split("|")[0].strip())
            row = df_live.loc[idx].copy()

            def txt(v):
                if pd.isna(v):
                    return ""
                return str(v)

            st.write("Modifier la ligne sélectionnée :")
            with st.form(key=skey("form_edit", str(idx))):
                r1c1, r1c2, r1c3 = st.columns([1.4,1.0,1.2])
                with r1c1:
                    st.markdown(f"**ID_Client :** {txt(row.get('ID_Client',''))}")
                with r1c2:
                    e_dossier = st.text_input("Dossier N", value=txt(row.get("Dossier N","")), key=skey("edit","dossier", str(idx)))
                with r1c3:
                    # IMPORTANT: use _date_or_none_safe(...) directly
                    e_date = st.date_input("Date (événement)", value=_date_or_none_safe(row.get("Date")), key=skey("edit","date", str(idx)))

                e_nom = st.text_input("Nom du client", value=txt(row.get("Nom","")), key=skey("edit","nom", str(idx)))

                r3c1, r3c2, r3c3 = st.columns([1.2,1.2,1.6])
                with r3c1:
                    if visa_categories:
                        edit_categories_options = visa_categories
                    else:
                        edit_categories_options = unique_nonempty(df_live["Categories"]) if "Categories" in df_live.columns else []
                    try:
                        init_cat_index = ([""]+edit_categories_options).index(txt(row.get("Categories",""))) if txt(row.get("Categories","")) in ([""]+edit_categories_options) else 0
                    except Exception:
                        init_cat_index = 0
                    e_cat = st.selectbox("Catégorie", options=[""]+edit_categories_options, index=init_cat_index, key=skey("edit","cat", str(idx)))
                with r3c2:
                    e_sub_options = []
                    if isinstance(e_cat, str) and e_cat.strip():
                        cat_key = canonical_key(e_cat)
                        if cat_key in visa_map_norm:
                            e_sub_options = visa_map_norm.get(cat_key, [])[:]
                        else:
                            if e_cat in visa_map:
                                e_sub_options = visa_map.get(e_cat, [])[:]
                    if not e_sub_options:
                        try:
                            e_sub_options = sorted({str(x).strip() for x in df_live["Sous-catégorie"].dropna().astype(str).tolist()})
                        except Exception:
                            e_sub_options = []
                    try:
                        init_sub_index = ([""]+e_sub_options).index(txt(row.get("Sous-categorie",""))) if txt(row.get("Sous-categorie","")) in ([""]+e_sub_options) else 0
                    except Exception:
                        init_sub_index = 0
                    e_sub = st.selectbox("Sous-catégorie", options=[""]+e_sub_options, index=init_sub_index, key=skey("edit","sub", str(idx)))
                with r3c3:
                    edit_specific = get_visa_options(e_cat, e_sub)
                    if edit_specific:
                        current = txt(row.get("Visa","")).strip()
                        options = [""] + edit_specific
                        try:
                            init_idx = options.index(current) if current in options else 0
                        except Exception:
                            init_idx = 0
                        e_visa = st.selectbox("Visa (options)", options=options, index=init_idx, key=skey("edit","visa", str(idx)))
                    else:
                        e_visa = st.text_input("Visa", value=txt(row.get("Visa","")), key=skey("edit","visa_text", str(idx)))

                r4c1, r4c2, r4c3, r4c4, r4c5 = st.columns([1.2,1.0,1.0,1.0,1.0])
                with r4c1:
                    e_montant = st.text_input("Montant honoraires (US $)", value=txt(row.get("Montant honoraires (US $)",0)), key=skey("edit","montant", str(idx)))
                with r4c2:
                    e_autres = st.text_input("Autres frais (US $)", value=txt(row.get("Autres frais (US $)",0)), key=skey("edit","autres", str(idx)))
                with r4c3:
                    try:
                        total_montant = _to_num(e_montant) + _to_num(e_autres)
                    except Exception:
                        total_montant = _to_num(row.get("Montant honoraires (US $)",0)) + _to_num(row.get("Autres frais (US $)",0))
                    st.text_input("Montant Total", value=str(total_montant), key=skey("edit","montant_total", str(idx)), disabled=True)
                with r4c4:
                    e_ac1 = st.text_input("Acompte 1", value=txt(row.get("Acompte 1",0)), key=skey("edit","ac1", str(idx)))
                with r4c5:
                    try:
                        paid_sum_preview = _to_num(e_ac1) + _to_num(row.get("Acompte 2",0)) + _to_num(row.get("Acompte 3",0)) + _to_num(row.get("Acompte 4",0))
                        solde_preview = total_montant - paid_sum_preview
                    except Exception:
                        solde_preview = row.get("Solde", 0)
                    st.text_input("Solde (calculé)", value=str(solde_preview), key=skey("edit","solde_preview", str(idx)), disabled=True)

                r5c1, r5c2, r5c3 = st.columns([1.0,1.0,1.0])
                with r5c1:
                    e_ac2 = st.text_input("Acompte 2", value=txt(row.get("Acompte 2",0)), key=skey("edit","ac2", str(idx)))
                with r5c2:
                    e_ac3 = st.text_input("Acompte 3", value=txt(row.get("Acompte 3",0)), key=skey("edit","ac3", str(idx)))
                with r5c3:
                    e_ac4 = st.text_input("Acompte 4", value=txt(row.get("Acompte 4",0)), key=skey("edit","ac4", str(idx)))

                r6c1, r6c2, r6c3 = st.columns([1.0,1.0,1.0])
                with r6c1:
                    e_ac2_date = st.date_input("Date Acompte 2", value=_date_or_none_safe(row.get("Date Acompte 2")), key=skey("edit","ac2_date", str(idx)))
                with r6c2:
                    e_ac3_date = st.date_input("Date Acompte 3", value=_date_or_none_safe(row.get("Date Acompte 3")), key=skey("edit","ac3_date", str(idx)))
                with r6c3:
                    e_ac4_date = st.date_input("Date Acompte 4", value=_date_or_none_safe(row.get("Date Acompte 4")), key=skey("edit","ac4_date", str(idx)))

                f1, f2, f3, f4, f5 = st.columns([1.0,1.0,1.0,1.0,0.6])
                with f1:
                    e_flag_envoye = st.checkbox("Dossiers envoyé", value=bool(int(row.get("Dossiers envoyé", 0))) if not pd.isna(row.get("Dossiers envoyé", 0)) else False, key=skey("edit","flag_envoye", str(idx)))
                with f2:
                    e_flag_approuve = st.checkbox("Dossier approuvé", value=bool(int(row.get("Dossier approuvé", 0))) if not pd.isna(row.get("Dossier approuvé", 0)) else False, key=skey("edit","flag_approuve", str(idx)))
                with f3:
                    e_flag_refuse = st.checkbox("Dossier refusé", value=bool(int(row.get("Dossier refusé", 0))) if not pd.isna(row.get("Dossier refusé", 0)) else False, key=skey("edit","flag_refuse", str(idx)))
                with f4:
                    e_flag_annule = st.checkbox("Dossier Annulé", value=bool(int(row.get("Dossier Annulé", 0))) if not pd.isna(row.get("Dossier Annulé", 0)) else False, key=skey("edit","flag_annule", str(idx)))
                with f5:
                    e_flag_rfe = st.checkbox("RFE", value=bool(int(row.get("RFE", 0))) if not pd.isna(row.get("RFE", 0)) else False, key=skey("edit","flag_rfe", str(idx)))
                d1, d2 = st.columns([1.6, 1.0])
                with d1:
                    e_flags_date = st.date_input("Date d'envoi / Date état", value=_date_or_none_safe(row.get("Date d'envoi")), key=skey("edit","flags_date", str(idx)))
                with d2:
                    st.markdown(" ")

                e_escrow = st.checkbox("Escrow", value=bool(int(row.get("Escrow", 0))) if not pd.isna(row.get("Escrow", 0)) else False, key=skey("edit","escrow", str(idx)))
                e_comments = st.text_area("Commentaires", value=txt(row.get("Commentaires","")), key=skey("edit","comments", str(idx)))

                save = st.form_submit_button("Enregistrer modifications")
                if save:
                    try:
                        df_live.at[idx, "Dossier N"] = e_dossier
                        df_live.at[idx, "Nom"] = e_nom
                        df_live.at[idx, "Date"] = pd.to_datetime(e_date)
                        df_live.at[idx, "Categories"] = e_cat
                        df_live.at[idx, "Sous-categorie"] = e_sub
                        df_live.at[idx, "Visa"] = e_visa
                        df_live.at[idx, "Montant honoraires (US $)"] = money_to_float(e_montant)
                        df_live.at[idx, "Autres frais (US $)"] = money_to_float(e_autres)
                        df_live.at[idx, "Acompte 1"] = money_to_float(e_ac1)
                        df_live.at[idx, "Acompte 2"] = money_to_float(e_ac2)
                        df_live.at[idx, "Acompte 3"] = money_to_float(e_ac3)
                        df_live.at[idx, "Acompte 4"] = money_to_float(e_ac4)
                        df_live.at[idx, "Date Acompte 2"] = pd.to_datetime(e_ac2_date) if e_ac2_date else pd.NaT
                        df_live.at[idx, "Date Acompte 3"] = pd.to_datetime(e_ac3_date) if e_ac3_date else pd.NaT
                        df_live.at[idx, "Date Acompte 4"] = pd.to_datetime(e_ac4_date) if e_ac4_date else pd.NaT
                        df_live.at[idx, "Escrow"] = 1 if e_escrow else 0
                        df_live.at[idx, "Dossiers envoyé"] = 1 if e_flag_envoye else 0
                        df_live.at[idx, "Dossier approuvé"] = 1 if e_flag_approuve else 0
                        df_live.at[idx, "Dossier refusé"] = 1 if e_flag_refuse else 0
                        df_live.at[idx, "Dossier Annulé"] = 1 if e_flag_annule else 0
                        df_live.at[idx, "RFE"] = 1 if e_flag_rfe else 0
                        df_live.at[idx, "Date d'envoi"] = pd.to_datetime(e_flags_date) if e_flags_date else pd.NaT
                        df_live.at[idx, "Dernière modification"] = datetime.now()
                        df_live.at[idx, "Modifié par"] = CURRENT_USER
                        df_live = recalc_payments_and_solde(df_live)
                        df_live.at[idx, "Solde à percevoir (US $)"] = df_live.at[idx, "Solde"]
                        df_live.at[idx, "Commentaires"] = e_comments
                        _set_df_live(df_live)
                        st.success("Modifications enregistrées.")
                    except Exception as e:
                        st.error(f"Erreur enregistrement: {e}")

    st.markdown("---")
    st.markdown("### Supprimer des dossiers")
    if df_live is None or df_live.empty:
        st.info("Aucun dossier à supprimer.")
    else:
        choices_del = [f"{i} | {df_live.at[i,'Dossier N'] if 'Dossier N' in df_live.columns else ''} | {df_live.at[i,'Nom'] if 'Nom' in df_live.columns else ''}" for i in range(len(df_live))]
        selected_to_del = st.multiselect("Sélectionnez les lignes à supprimer", options=choices_del, key=skey("del","select"))
        if st.button("Supprimer sélection"):
            if selected_to_del:
                idxs = [int(s.split("|")[0].strip()) for s in selected_to_del]
                try:
                    df_live = df_live.drop(index=idxs).reset_index(drop=True)
                    df_live = recalc_payments_and_solde(df_live)
                    _set_df_live(df_live)
                    st.success(f"{len(idxs)} ligne(s) supprimée(s).")
                except Exception as e:
                    st.error(f"Erreur suppression: {e}")
            else:
                st.warning("Aucune sélection pour suppression.")

# ---- Export tab ----
with tabs[5]:
    st.header("💾 Export")
    df_live = _get_df_live()
    if df_live is None or df_live.empty:
        st.info("Aucune donnée à exporter.")
    else:
        st.write(f"Vue en mémoire: {df_live.shape[0]} lignes, {df_live.shape[1]} colonnes")
        col1, col2 = st.columns(2)
        with col1:
            csv_bytes = df_live.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️ Export CSV", data=csv_bytes, file_name="Clients_export.csv", mime="text/csv")
        with col2:
            df_for_export = df_live.copy()
            try:
                montant_col = detect_montant_column(df_for_export) or "Montant honoraires (US $)"
                autres_col = detect_autres_column(df_for_export) or "Autres frais (US $)"
                acomptes_cols = detect_acompte_columns(df_for_export)
                df_for_export["_Montant_num_"] = df_for_export.get(montant_col,0).apply(lambda x: _to_num(x))
                df_for_export["_Autres_num_"] = df_for_export.get(autres_col,0).apply(lambda x: _to_num(x))
                for acc in acomptes_cols:
                    df_for_export[f"_num_{acc}"] = df_for_export.get(acc,0).apply(lambda x: _to_num(x))
                if acomptes_cols:
                    df_for_export["_Acomptes_sum_"] = df_for_export[[f"_num_{acc}" for acc in acomptes_cols]].sum(axis=1)
                else:
                    df_for_export["_Acomptes_sum_"] = 0.0
                df_for_export["Solde_formule"] = df_for_export["_Montant_num_"] + df_for_export["_Autres_num_"] - df_for_export["_Acomptes_sum_"]
                df_for_export["Solde à percevoir (US $)"] = df_for_export["Solde_formule"]
            except Exception:
                df_for_export["Solde_formule"] = df_for_export.get("Solde",0).apply(lambda x: _to_num(x))
                df_for_export["Solde à percevoir (US $)"] = df_for_export.get("Solde à percevoir (US $)",0).apply(lambda x: _to_num(x))
            drop_cols = [c for c in df_for_export.columns if c.startswith("_num_") or c in ["_Montant_num_","_Autres_num_","_Acomptes_sum_"]]
            try:
                df_export_final = df_for_export.drop(columns=drop_cols)
            except Exception:
                df_export_final = df_for_export.copy()
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_export_final.to_excel(writer, index=False, sheet_name="Clients")
            out_bytes = buf.getvalue()
            st.download_button("⬇️ Export XLSX (avec colonne Solde_formule)", data=out_bytes, file_name="Clients_export_with_Solde_formule.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.markdown("### Option avancée : XLSX avec formules (Payé & Solde)")
        if st.button("Générer XLSX avec formules Payé & Solde"):
            if not HAS_OPENPYXL:
                st.error("openpyxl non installé — impossible de générer le fichier avec formules.")
            else:
                buf2 = BytesIO()
                with pd.ExcelWriter(buf2, engine="openpyxl") as writer:
                    df_live.to_excel(writer, index=False, sheet_name="Clients")
                wb = load_workbook(filename=BytesIO(buf2.getvalue()))
                if "Clients" not in wb.sheetnames:
                    st.error("Feuille 'Clients' introuvable dans le workbook généré.")
                else:
                    ws = wb["Clients"]
                    headers = [cell.value for cell in ws[1]]
                    def col_letter_for(name: str):
                        try:
                            idx = headers.index(name) + 1
                            return get_column_letter(idx)
                        except Exception:
                            return None
                    col_paye = col_letter_for("Payé")
                    col_solde = col_letter_for("Solde")
                    col_solde_perc = col_letter_for("Solde à percevoir (US $)")
                    col_montant = col_letter_for("Montant honoraires (US $)")
                    col_autres = col_letter_for("Autres frais (US $)")
                    col_a1 = col_letter_for("Acompte 1")
                    col_a2 = col_letter_for("Acompte 2")
                    col_a3 = col_letter_for("Acompte 3")
                    col_a4 = col_letter_for("Acompte 4")
                    max_row = ws.max_row
                    if col_paye and any([col_a1,col_a2,col_a3,col_a4]):
                        for r in range(2, max_row+1):
                            parts = []
                            for c in (col_a1,col_a2,col_a3,col_a4):
                                if c:
                                    parts.append(f"{c}{r}")
                            if parts:
                                formula = "=IFERROR(" + "+".join(parts) + ",0)"
                                ws[f"{col_paye}{r}"] = formula
                    if col_solde and col_montant and col_autres and col_paye:
                        for r in range(2, max_row+1):
                            formula = f"=IFERROR({col_montant}{r}+{col_autres}{r}-{col_paye}{r},0)"
                            ws[f"{col_solde}{r}"] = formula
                            if col_solde_perc:
                                ws[f"{col_solde_perc}{r}"] = formula
                    out_buf = BytesIO()
                    wb.save(out_buf)
                    st.download_button("⬇️ Export XLSX (avec formules Payé & Solde)", data=out_buf.getvalue(), file_name="Clients_export_with_formulas.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# End of file
